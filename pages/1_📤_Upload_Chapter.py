import io
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from utils.column_detector import detect_columns
from utils.db import (
    get_all_chapter_numbers,
    chapter_exists,
    save_raw_chapter,
    rebuild_database,
    get_words_by_chapter,
)

st.set_page_config(page_title="Upload Chapter", page_icon="📤")
st.title("📤 নতুন Chapter Upload করো")

existing_chapters = []
try:
    existing_chapters = get_all_chapter_numbers()
except Exception:
    pass

suggested_next = (max(existing_chapters) + 1) if existing_chapters else 1
chapter_number = st.number_input("Chapter নাম্বার", min_value=1, step=1, value=suggested_next)

# --- Duplicate chapter warning ---
already_exists = False
try:
    already_exists = chapter_exists(chapter_number)
except Exception:
    pass

overwrite_confirmed = True
if already_exists:
    st.warning(
        f"⚠️ Chapter {chapter_number} আগে থেকেই database এ আছে। "
        "আবার upload করলে এই chapter এর আগের data মুছে গিয়ে নতুন data দিয়ে replace হয়ে যাবে।"
    )
    overwrite_confirmed = st.checkbox("হ্যাঁ, আমি জানি এবং Chapter টা overwrite করতে চাই")

uploaded_file = st.file_uploader("Excel বা CSV file আপলোড করো", type=["xlsx", "xls", "csv"])

if uploaded_file and (not already_exists or overwrite_confirmed):
    if uploaded_file.name.endswith(".csv"):
        raw_df = pd.read_csv(uploaded_file)
    else:
        raw_df = pd.read_excel(uploaded_file)

    st.write("**Preview (প্রথম ৫ row):**")
    st.dataframe(raw_df.head())

    korean_guess, bangla_guess, notes = detect_columns(raw_df)
    if notes:
        for n in notes:
            st.info(n)

    cols = list(raw_df.columns)
    korean_col = st.selectbox(
        "Korean word এর column বেছে নাও",
        cols,
        index=cols.index(korean_guess) if korean_guess in cols else 0,
    )
    bangla_col = st.selectbox(
        "Bangla meaning এর column বেছে নাও",
        cols,
        index=cols.index(bangla_guess) if bangla_guess in cols else (1 if len(cols) > 1 else 0),
    )

    if st.button("💾 Save করো ও Unique Word বের করো", type="primary"):
        # basic cleaning: strip + drop empty korean rows
        clean = raw_df[[korean_col, bangla_col]].copy()
        clean.columns = ["Korean", "Bangla"]
        clean["Korean"] = clean["Korean"].astype(str).str.strip()
        clean["Bangla"] = clean["Bangla"].astype(str).str.strip()
        clean = clean[(clean["Korean"] != "") & (clean["Korean"].str.lower() != "nan")]

        pairs = list(zip(clean["Korean"], clean["Bangla"]))

        with st.spinner("Database এ raw data save হচ্ছে এবং পুরো hisab rebuild করা হচ্ছে..."):
            save_raw_chapter(chapter_number, pairs)
            rebuild_database()
            final_rows = get_words_by_chapter(chapter_number)

        final_df = pd.DataFrame(final_rows)
        st.session_state["final_df"] = final_df
        st.session_state["saved_chapter_number"] = chapter_number

    if "final_df" in st.session_state and st.session_state.get("saved_chapter_number") == chapter_number:
        final_df = st.session_state["final_df"]

        if final_df.empty:
            st.warning("এই Chapter এ কোনো নতুন/unique word পাওয়া যায়নি (সব word আগের অন্য chapter এ আগেই আছে)।")
        else:
            st.success(f"Chapter {chapter_number} এ মোট {len(final_df)} টা সঠিক/চূড়ান্ত unique word আছে (rebuild এর পর)।")
            st.dataframe(final_df[["korean_word", "bangla_meaning"]])

            csv_bytes = final_df[["korean_word", "bangla_meaning"]].to_csv(
                index=False, header=["Korean word", "Bangla word"]
            ).encode("utf-8-sig")

            wb = Workbook()
            ws = wb.active
            ws.title = f"Chapter_{chapter_number}"
            ws.append(["Korean word", "Bangla word"])
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            for _, row in final_df.iterrows():
                ws.append([row["korean_word"], row["bangla_meaning"]])
            ws.column_dimensions["A"].width = 30
            ws.column_dimensions["B"].width = 30

            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)

            col1, col2 = st.columns(2)
            col1.download_button(
                "⬇️ Excel Download",
                data=excel_buffer,
                file_name=f"chapter_{chapter_number}_unique.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            col2.download_button(
                "⬇️ CSV Download (2 column)",
                data=csv_bytes,
                file_name=f"chapter_{chapter_number}_unique.csv",
                mime="text/csv",
            )

        st.caption(
            "নোট: এই ফলাফল সব chapter এর মধ্যে chapter-number অনুযায়ী ক্রম মেনে হিসাব করে বের করা, "
            "upload করার ক্রম অনুযায়ী না। তাই পরে মাঝের কোনো chapter upload করলে অন্য chapter গুলোর "
            "সংখ্যাও সেই অনুযায়ী সঠিকভাবে বদলে যেতে পারে।"
        )
