import io
import json
import re
import time

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from utils.db import (
    get_all_chapter_numbers,
    chapter_exists,
    save_raw_chapter,
    rebuild_database,
    get_words_by_chapter,
)
from utils.auth import is_admin, require_admin

st.set_page_config(page_title="Image → Excel Extractor", page_icon="🖼️", layout="wide")
st.title("🖼️ ছবি থেকে Excel (Korean-Bangla Vocab Table)")
st.caption(
    "স্ক্যান করা vocabulary table এর ছবি upload করো — Gemini Vision দিয়ে ক্রম নম্বর, "
    "Korean word ও Bangla অর্থ বের করে Excel বানিয়ে দেবে।"
)

EXTRACT_PROMPT = """এই ছবির টেবিল থেকে প্রতিটা সারির ক্রম/serial নম্বর, কোরিয়ান শব্দ, এবং তার বাংলা অর্থ
নির্ভুলভাবে বের করে শুধু JSON array আকারে দাও। ফরম্যাট:
[{"serial": "...", "korean": "...", "bangla": "..."}]
কোনো ব্যাখ্যা বা অতিরিক্ত টেক্সট দিও না, শুধু JSON। অনুমান করে কিছু বানিও না, যা স্পষ্ট পড়া যাচ্ছে না
সেটা "?" দিয়ে চিহ্নিত করো।"""

# আমাদের অন্য AI feature গুলোর মতোই একই candidate model list (কোনো একটা
# deprecate হয়ে গেলেও app যাতে immediately না ভাঙে)
CANDIDATE_MODELS = ["gemini-3.1-flash-lite", "gemini-2.5-flash", "gemini-flash-latest"]


def _parse_json_array(text):
    text = text.strip()
    text = re.sub(r"^```json\s*|\s*```$", "", text.strip())
    match = re.search(r"\[.*\]", text, re.DOTALL)
    if match:
        text = match.group(0)
    return json.loads(text)


def extract_from_image(image_bytes, mime_type, api_key, model_name, max_retries=3):
    """একটা image Gemini Vision এ পাঠিয়ে JSON array হিসেবে rows ফেরত দেয়।
    429 (rate limit) এ পড়লে exponential backoff দিয়ে retry করে।"""
    from google import genai as google_genai
    from google.genai import types

    client = google_genai.Client(api_key=api_key)
    image_part = types.Part.from_bytes(data=image_bytes, mime_type=mime_type)

    last_err = None
    for attempt in range(max_retries):
        try:
            response = client.models.generate_content(
                model=model_name, contents=[image_part, EXTRACT_PROMPT]
            )
            return _parse_json_array(response.text)
        except Exception as e:
            last_err = e
            err_str = str(e)
            if "429" in err_str or "RESOURCE_EXHAUSTED" in err_str:
                wait = 2 ** (attempt + 1)  # 2s, 4s, 8s
                time.sleep(wait)
                continue
            else:
                raise
    raise last_err


def extract_with_model_fallback(image_bytes, mime_type, api_key):
    """candidate model list ক্রমান্বয়ে চেষ্টা করে, একটা কাজ না করলে পরেরটা।"""
    last_err = None
    for model_name in CANDIDATE_MODELS:
        try:
            return extract_from_image(image_bytes, mime_type, api_key, model_name), model_name
        except Exception as e:
            last_err = e
    raise last_err


# ---------- UI ----------

api_key = st.secrets.get("GEMINI_API_KEY")
if not api_key:
    st.error(
        "⚠️ GEMINI_API_KEY Streamlit secrets এ পাওয়া যায়নি। "
        "Settings → Secrets এ গিয়ে GEMINI_API_KEY যোগ করো (Paragraph Translator এর জন্য যেটা ব্যবহার করেছ, সেটাই)।"
    )
    st.stop()

uploaded_files = st.file_uploader(
    "একটা বা একাধিক ছবি আপলোড করো (JPG/PNG)",
    type=["jpg", "jpeg", "png"],
    accept_multiple_files=True,
)

if uploaded_files:
    st.write(f"**{len(uploaded_files)} টা ছবি আপলোড হয়েছে**")

    if st.button("🔍 সব ছবি Process করো", type="primary"):
        all_rows = []
        errors = []
        progress = st.progress(0, text="শুরু হচ্ছে...")

        for i, file in enumerate(uploaded_files):
            progress.progress((i) / len(uploaded_files), text=f"Processing: {file.name}")
            image_bytes = file.read()
            mime_type = file.type or "image/jpeg"

            try:
                rows, used_model = extract_with_model_fallback(image_bytes, mime_type, api_key)
                for r in rows:
                    all_rows.append(
                        {
                            "Serial No": r.get("serial", "?"),
                            "Korean Word": r.get("korean", "?"),
                            "Bangla Meaning": r.get("bangla", "?"),
                            "Source Image": file.name,
                        }
                    )
            except Exception as e:
                errors.append(f"❌ **{file.name}** প্রসেস করা যায়নি: {e}")
                continue  # বাকি ছবি প্রসেস করা চলতে থাকবে

        progress.progress(1.0, text="সম্পন্ন ✅")

        if errors:
            st.warning("কিছু ছবিতে সমস্যা হয়েছে (বাকি গুলো ঠিকভাবে প্রসেস হয়েছে):")
            for e in errors:
                st.write(e)

        if all_rows:
            df = pd.DataFrame(all_rows)
            st.session_state["extracted_df"] = df
        elif not errors:
            st.info("কোনো data পাওয়া যায়নি।")

if "extracted_df" in st.session_state:
    df = st.session_state["extracted_df"]
    st.divider()
    st.subheader(f"ফলাফল — মোট {len(df)} টা row (Download এর আগে চোখ বুলিয়ে ভুল ঠিক করো)")

    edited_df = st.data_editor(df, use_container_width=True, num_rows="dynamic")
    st.session_state["extracted_df"] = edited_df

    # ---------- Excel বানানো ----------
    export_df = edited_df[["Serial No", "Korean Word", "Bangla Meaning"]]

    wb = Workbook()
    ws = wb.active
    ws.title = "Extracted Vocab"
    ws.append(["Serial No", "Korean Word", "Bangla Meaning"])
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for _, row in export_df.iterrows():
        ws.append([row["Serial No"], row["Korean Word"], row["Bangla Meaning"]])
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 30

    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    st.download_button(
        "⬇️ Excel Download করো",
        data=excel_buffer,
        file_name="extracted_vocab.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    if st.button("🗑️ ফলাফল Clear করো"):
        del st.session_state["extracted_df"]
        st.rerun()

    st.divider()
    st.subheader("📥 Chapter Database এ Save করো")
    st.caption(
        "এই word গুলো সরাসরি আমাদের chapter-wise vocabulary database এ যোগ করতে পারো "
        "(ঠিক Upload Chapter page যেভাবে কাজ করে, সেভাবেই — আগের chapter এর সাথে duplicate check হবে)।"
    )

    if not is_admin():
        st.info(
            "🔒 Database এ save করার জন্য Admin login লাগবে। এই section দেখা যাচ্ছে না — "
            "Admin হলে বাম sidebar থেকে যেকোনো Manager page (যেমন Upload Chapter) এ গিয়ে login করো, "
            "তারপর এই page এ ফিরে এলে save option চলে আসবে।"
        )
    else:

        try:
            existing_chapters = get_all_chapter_numbers()
        except Exception:
            existing_chapters = []
        suggested_next = (max(existing_chapters) + 1) if existing_chapters else 1

        save_chapter_number = st.number_input(
            "কোন Chapter নাম্বারে Save করবে?", min_value=1, step=1, value=suggested_next, key="save_chapter_num"
        )

        already_exists = False
        try:
            already_exists = chapter_exists(save_chapter_number)
        except Exception:
            pass

        overwrite_confirmed = True
        if already_exists:
            st.warning(
                f"⚠️ Chapter {save_chapter_number} আগে থেকেই database এ আছে। "
                "Save করলে এই chapter এর আগের data মুছে গিয়ে নতুন data দিয়ে replace হয়ে যাবে।"
            )
            overwrite_confirmed = st.checkbox(
                f"হ্যাঁ, Chapter {save_chapter_number} overwrite করতে চাই", key="overwrite_confirm_img"
            )

        if st.button(
            f"💾 Chapter {save_chapter_number} এ Save করো",
            type="primary",
            disabled=already_exists and not overwrite_confirmed,
        ):
            clean = edited_df[["Korean Word", "Bangla Meaning"]].copy()
            clean.columns = ["Korean", "Bangla"]
            clean["Korean"] = clean["Korean"].astype(str).str.strip()
            clean["Bangla"] = clean["Bangla"].astype(str).str.strip()
            clean = clean[(clean["Korean"] != "") & (clean["Korean"].str.lower() != "nan")]
            pairs = list(zip(clean["Korean"], clean["Bangla"]))

            with st.spinner("Database এ save হচ্ছে এবং পুরো hisab rebuild করা হচ্ছে..."):
                save_raw_chapter(save_chapter_number, pairs)
                rebuild_database()
                final_rows = get_words_by_chapter(save_chapter_number)

            final_df = pd.DataFrame(final_rows)
            if final_df.empty:
                st.warning("Save হয়েছে, কিন্তু এই chapter এ কোনো নতুন/unique word নেই (সব word আগের chapter এ আগেই আছে)।")
            else:
                st.success(f"✅ Chapter {save_chapter_number} এ {len(final_df)} টা unique word save হয়ে গেছে।")
                st.dataframe(final_df[["korean_word", "bangla_meaning"]], use_container_width=True)
            st.caption("সম্পূর্ণ chapter দেখতে বা আরও download করতে **📂 Browse History** page এ যাও।")
